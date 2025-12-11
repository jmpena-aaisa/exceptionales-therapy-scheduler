from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook

from .time_utils import BLOCKS, DAY_ORDER, block_to_range, range_to_block


@dataclass
class Session:
    therapist_id: str
    room_id: str
    day: str
    block: int
    specialty: str
    patients: List[str] = field(default_factory=list)

    @property
    def size(self) -> int:
        return len(self.patients)


def aggregate_sessions(schedule: Iterable[Dict[str, str]]) -> List[Session]:
    grouped: Dict[Tuple[str, str, str, int, str], Session] = {}
    for item in schedule:
        block = range_to_block(item["time"])
        key = (
            item["therapist_id"],
            item["room_id"],
            item["day"],
            block,
            item["specialty"],
        )
        if key not in grouped:
            grouped[key] = Session(
                therapist_id=item["therapist_id"],
                room_id=item["room_id"],
                day=item["day"],
                block=block,
                specialty=item["specialty"],
            )
        grouped[key].patients.append(item["patient_id"])
    return list(grouped.values())


def _render_cell(session: Session) -> str:
    patients = ", ".join(sorted(session.patients))
    return f"{session.specialty} | {patients} | {session.therapist_id} | {session.room_id} | n={session.size}"


def export_excel(schedule: List[Dict[str, str]], output_path: Path) -> None:
    sessions = aggregate_sessions(schedule)
    wb = Workbook()
    wb.remove(wb.active)

    _add_room_tabs(wb, sessions)
    _add_therapist_tab(wb, sessions)
    _add_patient_tab(wb, sessions)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def _add_room_tabs(wb: Workbook, sessions: List[Session]) -> None:
    sessions_by_room: Dict[str, Dict[Tuple[str, int], Session]] = defaultdict(dict)
    for s in sessions:
        sessions_by_room[s.room_id][(s.day, s.block)] = s

    for room_id in sorted(sessions_by_room.keys()):
        ws = wb.create_sheet(title=room_id[:31])
        _write_header(ws)
        for row_idx, block in enumerate(BLOCKS, start=2):
            ws.cell(row=row_idx, column=1, value=block_to_range(block))
            for col_idx, day in enumerate(DAY_ORDER, start=2):
                session = sessions_by_room[room_id].get((day, block))
                if session:
                    ws.cell(row=row_idx, column=col_idx, value=_render_cell(session))
        _autosize(ws, len(DAY_ORDER) + 1)


def _add_therapist_tab(wb: Workbook, sessions: List[Session]) -> None:
    ws = wb.create_sheet(title="Therapists")
    therapist_ids = sorted({s.therapist_id for s in sessions})
    ws.cell(row=1, column=1, value="Day")
    ws.cell(row=1, column=2, value="Time")
    for idx, tid in enumerate(therapist_ids, start=3):
        ws.cell(row=1, column=idx, value=tid)

    row_idx = 2
    for day in DAY_ORDER:
        for block in BLOCKS:
            ws.cell(row=row_idx, column=1, value=day)
            ws.cell(row=row_idx, column=2, value=block_to_range(block))
            for col_idx, tid in enumerate(therapist_ids, start=3):
                session = next(
                    (
                        s
                        for s in sessions
                        if s.day == day and s.block == block and s.therapist_id == tid
                    ),
                    None,
                )
                if session:
                    ws.cell(row=row_idx, column=col_idx, value=_render_cell(session))
            row_idx += 1
    _autosize(ws, len(therapist_ids) + 2)


def _add_patient_tab(wb: Workbook, sessions: List[Session]) -> None:
    ws = wb.create_sheet(title="Patients")
    patient_ids = sorted({p for s in sessions for p in s.patients})
    ws.cell(row=1, column=1, value="Day")
    ws.cell(row=1, column=2, value="Time")
    for idx, pid in enumerate(patient_ids, start=3):
        ws.cell(row=1, column=idx, value=pid)

    row_idx = 2
    for day in DAY_ORDER:
        for block in BLOCKS:
            ws.cell(row=row_idx, column=1, value=day)
            ws.cell(row=row_idx, column=2, value=block_to_range(block))
            for col_idx, pid in enumerate(patient_ids, start=3):
                session = next(
                    (
                        s
                        for s in sessions
                        if s.day == day and s.block == block and pid in s.patients
                    ),
                    None,
                )
                if session:
                    ws.cell(row=row_idx, column=col_idx, value=_render_cell(session))
            row_idx += 1
    _autosize(ws, len(patient_ids) + 2)


def _write_header(ws) -> None:
    ws.cell(row=1, column=1, value="Time")
    for idx, day in enumerate(DAY_ORDER, start=2):
        ws.cell(row=1, column=idx, value=day)


def _autosize(ws, num_columns: int) -> None:
    for col in range(1, num_columns + 1):
        max_len = 0
        col_letter = ws.cell(row=1, column=col).column_letter
        for cell in ws[col_letter]:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 80)
